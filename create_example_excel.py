#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
建立 Excel 範例檔案
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta

# 建立新的工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "時程規劃"

# 設定標題列
headers = [
    "環境名稱",
    "環境目的",
    "執行梯次",
    "驗證起日",
    "驗證迄日",
    "工作內容",
    "狀態",
    "中介檔",
    "資料基準日",
    "京城封版日",
    "京城傳送中介檔日",
    "備注說明"
]

# 寫入標題列
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, size=12)
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 設定列高
ws.row_dimensions[1].height = 25

# 資料範例
data = [
    # 資轉驗證環境
    {
        "環境名稱": "資轉驗證環境",
        "環境目的": "資料轉換驗證與測試",
        "執行梯次": "第一梯次",
        "驗證起日": datetime(2024, 1, 15),
        "驗證迄日": datetime(2024, 1, 20),
        "工作內容": "IT前置準備之1.永豐BSP確認接收日",
        "狀態": "進行中",
        "中介檔": "BSP_20240115.xlsx",
        "資料基準日": datetime(2024, 1, 10),
        "京城封版日": datetime(2024, 1, 12),
        "京城傳送中介檔日": datetime(2024, 1, 14),
        "備注說明": "需確認資料完整性"
    },
    {
        "環境名稱": "資轉驗證環境",
        "環境目的": "資料轉換驗證與測試",
        "執行梯次": "第一梯次",
        "驗證起日": datetime(2024, 1, 15),
        "驗證迄日": datetime(2024, 1, 20),
        "工作內容": "IT前置準備之2.永豐BSP DB倒檔",
        "狀態": "進行中",
        "中介檔": "BSP_DB_20240115.dump",
        "資料基準日": datetime(2024, 1, 10),
        "京城封版日": datetime(2024, 1, 12),
        "京城傳送中介檔日": datetime(2024, 1, 14),
        "備注說明": "需確認DB版本"
    },
    {
        "環境名稱": "資轉驗證環境",
        "環境目的": "資料轉換驗證與測試",
        "執行梯次": "第一梯次",
        "驗證起日": datetime(2024, 1, 15),
        "驗證迄日": datetime(2024, 1, 20),
        "工作內容": "IT前置準備之3.永豐BSP AP確認",
        "狀態": "進行中",
        "中介檔": "BSP_AP_20240115.zip",
        "資料基準日": datetime(2024, 1, 10),
        "京城封版日": datetime(2024, 1, 12),
        "京城傳送中介檔日": datetime(2024, 1, 14),
        "備注說明": "AP版本需與DB一致"
    },
    {
        "環境名稱": "資轉驗證環境",
        "環境目的": "資料轉換驗證與測試",
        "執行梯次": "第二梯次",
        "驗證起日": datetime(2024, 1, 22),
        "驗證迄日": datetime(2024, 1, 27),
        "工作內容": "IT前置準備之1.永豐BSP確認接收日",
        "狀態": "待開始",
        "中介檔": "BSP_20240122.xlsx",
        "資料基準日": datetime(2024, 1, 18),
        "京城封版日": datetime(2024, 1, 20),
        "京城傳送中介檔日": datetime(2024, 1, 21),
        "備注說明": "第二梯次資料"
    },
    {
        "環境名稱": "資轉驗證環境",
        "環境目的": "資料轉換驗證與測試",
        "執行梯次": "第二梯次",
        "驗證起日": datetime(2024, 1, 22),
        "驗證迄日": datetime(2024, 1, 27),
        "工作內容": "IT前置準備之2.永豐BSP DB倒檔",
        "狀態": "待開始",
        "中介檔": "BSP_DB_20240122.dump",
        "資料基準日": datetime(2024, 1, 18),
        "京城封版日": datetime(2024, 1, 20),
        "京城傳送中介檔日": datetime(2024, 1, 21),
        "備注說明": "需確認DB版本一致性"
    },
    {
        "環境名稱": "資轉驗證環境",
        "環境目的": "資料轉換驗證與測試",
        "執行梯次": "第二梯次",
        "驗證起日": datetime(2024, 1, 22),
        "驗證迄日": datetime(2024, 1, 27),
        "工作內容": "IT前置準備之3.永豐BSP AP確認",
        "狀態": "待開始",
        "中介檔": "BSP_AP_20240122.zip",
        "資料基準日": datetime(2024, 1, 18),
        "京城封版日": datetime(2024, 1, 20),
        "京城傳送中介檔日": datetime(2024, 1, 21),
        "備注說明": "AP需與第一梯次版本一致"
    },
    # 平測切轉環境
    {
        "環境名稱": "平測切轉環境",
        "環境目的": "平行測試與系統切換驗證",
        "執行梯次": "第一梯次",
        "驗證起日": datetime(2024, 2, 1),
        "驗證迄日": datetime(2024, 2, 5),
        "工作內容": "IT前置準備之1.永豐BSP確認接收日",
        "狀態": "待開始",
        "中介檔": "BSP_20240201.xlsx",
        "資料基準日": datetime(2024, 1, 28),
        "京城封版日": datetime(2024, 1, 30),
        "京城傳送中介檔日": datetime(2024, 1, 31),
        "備注說明": "平測環境首次資料"
    },
    {
        "環境名稱": "平測切轉環境",
        "環境目的": "平行測試與系統切換驗證",
        "執行梯次": "第一梯次",
        "驗證起日": datetime(2024, 2, 1),
        "驗證迄日": datetime(2024, 2, 5),
        "工作內容": "IT前置準備之2.永豐BSP DB倒檔",
        "狀態": "待開始",
        "中介檔": "BSP_DB_20240201.dump",
        "資料基準日": datetime(2024, 1, 28),
        "京城封版日": datetime(2024, 1, 30),
        "京城傳送中介檔日": datetime(2024, 1, 31),
        "備注說明": "DB需完整備份"
    },
    {
        "環境名稱": "平測切轉環境",
        "環境目的": "平行測試與系統切換驗證",
        "執行梯次": "第一梯次",
        "驗證起日": datetime(2024, 2, 1),
        "驗證迄日": datetime(2024, 2, 5),
        "工作內容": "IT前置準備之3.永豐BSP AP確認",
        "狀態": "待開始",
        "中介檔": "BSP_AP_20240201.zip",
        "資料基準日": datetime(2024, 1, 28),
        "京城封版日": datetime(2024, 1, 30),
        "京城傳送中介檔日": datetime(2024, 1, 31),
        "備注說明": "AP版本需與資轉環境一致"
    },
    {
        "環境名稱": "平測切轉環境",
        "環境目的": "平行測試與系統切換驗證",
        "執行梯次": "第二梯次",
        "驗證起日": datetime(2024, 2, 8),
        "驗證迄日": datetime(2024, 2, 12),
        "工作內容": "IT前置準備之1.永豐BSP確認接收日",
        "狀態": "待開始",
        "中介檔": "BSP_20240208.xlsx",
        "資料基準日": datetime(2024, 2, 5),
        "京城封版日": datetime(2024, 2, 7),
        "京城傳送中介檔日": datetime(2024, 2, 7),
        "備注說明": "第二梯次資料更新"
    },
    {
        "環境名稱": "平測切轉環境",
        "環境目的": "平行測試與系統切換驗證",
        "執行梯次": "第二梯次",
        "驗證起日": datetime(2024, 2, 8),
        "驗證迄日": datetime(2024, 2, 12),
        "工作內容": "IT前置準備之2.永豐BSP DB倒檔",
        "狀態": "待開始",
        "中介檔": "BSP_DB_20240208.dump",
        "資料基準日": datetime(2024, 2, 5),
        "京城封版日": datetime(2024, 2, 7),
        "京城傳送中介檔日": datetime(2024, 2, 7),
        "備注說明": "DB增量更新"
    },
    {
        "環境名稱": "平測切轉環境",
        "環境目的": "平行測試與系統切換驗證",
        "執行梯次": "第二梯次",
        "驗證起日": datetime(2024, 2, 8),
        "驗證迄日": datetime(2024, 2, 12),
        "工作內容": "IT前置準備之3.永豐BSP AP確認",
        "狀態": "待開始",
        "中介檔": "BSP_AP_20240208.zip",
        "資料基準日": datetime(2024, 2, 5),
        "京城封版日": datetime(2024, 2, 7),
        "京城傳送中介檔日": datetime(2024, 2, 7),
        "備注說明": "AP版本確認"
    },
]

# 寫入資料
for row_idx, row_data in enumerate(data, start=2):
    for col_idx, header in enumerate(headers, start=1):
        value = row_data.get(header, "")
        
        # 日期格式處理
        if isinstance(value, datetime):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.number_format = "YYYY-MM-DD"
        else:
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 設定對齊
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 自動調整欄寬
column_widths = {
    "A": 20,  # 環境名稱
    "B": 25,  # 環境目的
    "C": 15,  # 執行梯次
    "D": 15,  # 驗證起日
    "E": 15,  # 驗證迄日
    "F": 40,  # 工作內容
    "G": 12,  # 狀態
    "H": 25,  # 中介檔
    "I": 15,  # 資料基準日
    "J": 15,  # 京城封版日
    "K": 20,  # 京城傳送中介檔日
    "L": 30,  # 備注說明
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# 設定列高（除了標題列）
for row in range(2, len(data) + 2):
    ws.row_dimensions[row].height = 20

# 儲存檔案
output_file = "resource/範例_藍圖之對應時程環境規劃.xlsx"
wb.save(output_file)
print(f"Excel 範例檔案已建立：{output_file}")
print(f"共 {len(data)} 筆資料")

